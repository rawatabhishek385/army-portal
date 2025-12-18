# registration/admin.py
import csv
import json
from datetime import timedelta
import zipfile
import os as _os  # for urandom

from django import forms
from django.conf import settings
from django.contrib import admin
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.urls import reverse, path
from django.utils import timezone
from django.utils.html import format_html
import openpyxl
from openpyxl.utils import get_column_letter

from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

from .models import CandidateProfile
from results.models import CandidateAnswer
from questions.models import QuestionPaper


# -------------------------
# Custom Admin Form with Validation
# -------------------------
class CandidateProfileAdminForm(forms.ModelForm):
    class Meta:
        model = CandidateProfile
        fields = "__all__"

    def clean(self):
        cleaned_data = super().clean()

        # Update instance with form data before validation
        instance = self.instance
        for field_name in [
            "trade",
            "primary_practical_marks",
            "primary_viva_marks",
            # "secondary_practical_marks",
            # "secondary_viva_marks",
        ]:
            if field_name in cleaned_data:
                setattr(instance, field_name, cleaned_data[field_name])

        # Run model validation
        try:
            instance.full_clean()
        except ValidationError as e:
            if hasattr(e, "message_dict"):
                for field, messages in e.message_dict.items():
                    if isinstance(messages, list):
                        for message in messages:
                            self.add_error(field, message)
                    else:
                        self.add_error(field, messages)
            else:
                raise forms.ValidationError(str(e))

        return cleaned_data


# -------------------------
# CSV exporter (candidate answers)
# -------------------------
def export_candidate_answers(modeladmin, request, queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="selected_candidates_answers.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "Army Number",
            "Candidate Name",
            "Paper Title",
            "Question ID",
            "Question Text",
            "Answer",
            "Category",
            "Submitted At",
        ]
    )

    answers = (
        CandidateAnswer.objects.filter(candidate__in=queryset)
        .select_related("candidate", "paper", "question")
        .order_by("candidate__id", "paper_id", "question_id")
    )

    for ans in answers:
        writer.writerow(
            [
                getattr(ans.candidate, "army_no", ""),
                ans.candidate.name if ans.candidate else "",
                getattr(ans.paper, "title", ""),
                getattr(ans.question, "id", ""),
                getattr(ans.question, "text", ""),
                getattr(ans, "answer", ""),
                ans.effective_category,  # computed: primary/secondary
                getattr(ans, "submitted_at", ""),
            ]
        )
    return response


export_candidate_answers.short_description = "Export selected candidates' answers to CSV"


# -------------------------
# Excel exporter (candidates)
# -------------------------
def export_candidates_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    columns = [
        "Army No",
        "Rank",
        "Name",
        "Photo",
        "Trade",
        "DOB",
        "Father Name",
        "Date of Enrolment",
        "Aadhar Number",
        "Training Center",
        "District",
        "State",
        "Primary Qualification",
        "Primary Duration",
        "Primary Credits",
        # "Secondary Qualification",
        # "Secondary Duration",
        # "Secondary Credits",
        "NSQF Level",
        "Exam Center",
        "Shift",
        "Created At",
    ]

    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num).value = column_title

    for row_num, candidate in enumerate(queryset, 2):
        # Safe access for optional fields
        photo = getattr(candidate, "photograph", None)
        photo_url = getattr(photo, "url", "") if photo else ""
        father_name = getattr(candidate, "father_name", "")
        aadhar_number = getattr(candidate, "aadhar_number", "")

        data = [
            candidate.army_no,
            candidate.rank,
            candidate.name,
            photo_url,
            getattr(candidate.trade, "name", str(candidate.trade)) if getattr(candidate, "trade", None) else "",
            candidate.dob,
            father_name,
            candidate.doe.strftime("%Y-%m-%d") if candidate.doe else "",
            aadhar_number,
            getattr(candidate, "training_center", ""),
            getattr(candidate, "district", ""),
            getattr(candidate, "state", ""),
            candidate.primary_qualification,
            candidate.primary_duration,
            candidate.primary_credits,
            # candidate.secondary_qualification,
            # candidate.secondary_duration,
            # candidate.secondary_credits,
            candidate.nsqf_level,
            candidate.exam_center,
            str(candidate.shift) if candidate.shift else "",
            candidate.created_at.strftime("%Y-%m-%d %H:%M") if candidate.created_at else "",
        ]
        for col_num, cell_value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="candidates.xlsx"'
    wb.save(response)
    return response


export_candidates_excel.short_description = "Export selected candidates to Excel"


# -------------------------
# Helper: Build a multi-sheet workbook for .dat payload
# -------------------------
def _build_export_workbook(queryset):
    from openpyxl import Workbook
    from io import BytesIO

    # local imports to avoid circular import at module level
    from questions.models import QuestionPaper
    from results.models import CandidateAnswer
    from questions.models import ExamSession

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "S.No",
        "Name",
        "Center",
        "Photo",
        "Fathers_Name",
        "DOB",
        "Rank",
        "Trade",
        "Army_No",
        "Adhaar_No",
        "Primary Qualification",
        "Primary Duration",
        "Primary Credits",
        # "Secondary Qualification",
        # "Secondary Duration",
        # "Secondary Credits",
        "NSQF Level",
        "Training_Center",
        "District",
        "State",
        "Viva_1",
        "Viva_2",
        "Practical_1",
        "Practical_2",
        "Army_No",
        "Exam_Type",
        "Part",
        "Question",
        "Answer",
        "Correct_Answer",
        "Max_Marks",
    ]

    ws.append(headers)
    serial = 1

    for candidate in queryset:
        # Safe access for optional fields
        photo = getattr(candidate, "photograph", None)
        photo_url = getattr(photo, "url", "") if photo else ""
        father_name = getattr(candidate, "father_name", "")
        aadhar_number = getattr(candidate, "aadhar_number", "")

        # For each candidate, find their exam sessions (these contain the assigned questions)
        sessions = (
            ExamSession.objects
            .filter(user=candidate.user)
            .select_related("paper")
            .prefetch_related("examquestion_set__question")
            .order_by("-started_at")
        )

        # If candidate has no sessions, optionally fallback to candidate_answers to include any orphan answers
        if not sessions.exists():
            # Optional: include papers where candidate actually has answers (helps if sessions were deleted)
            # We'll iterate papers that have CandidateAnswer rows for this candidate.
            papers_with_answers = QuestionPaper.objects.filter(candidate_answers__candidate=candidate).distinct()
            for paper in papers_with_answers:
                questions = paper.questions.all().order_by("id")
                for q in questions:
                    ans = CandidateAnswer.objects.filter(candidate=candidate, paper=paper, question=q).first()
                    row = [
                        serial,
                        candidate.name,
                        candidate.exam_center,
                        photo_url,
                        father_name,
                        candidate.dob,
                        candidate.rank,
                        candidate.trade.name if getattr(candidate, "trade", None) else "",
                        candidate.army_no,
                        aadhar_number,
                        candidate.primary_qualification,
                        candidate.primary_duration,
                        candidate.primary_credits,
                        candidate.secondary_qualification,
                        candidate.secondary_duration,
                        candidate.secondary_credits,
                        candidate.nsqf_level,
                        getattr(candidate, "training_center", ""),
                        getattr(candidate, "district", ""),
                        getattr(candidate, "state", ""),
                        candidate.primary_viva_marks,
                        # candidate.secondary_viva_marks,
                        candidate.primary_practical_marks,
                        # candidate.secondary_practical_marks,
                        candidate.army_no,
                        "Secondary" if getattr(paper, "is_common", False) else "Primary",
                        q.part,
                        q.text,
                        ans.answer if ans and ans.answer is not None else "N/A",
                        getattr(q, "correct_answer", None),
                        q.marks if hasattr(q, "marks") else None,
                    ]
                    ws.append(row)
                    serial += 1
            continue

        # Otherwise iterate sessions and the ExamQuestions (assigned questions)
        for session in sessions:
            paper = session.paper  # may be None if paper row deleted; handle gracefully
            exam_questions = session.questions  # property returns ordered ExamQuestion queryset

            for eq in exam_questions:
                q = eq.question
                # Try to fetch CandidateAnswer by candidate + paper + question.
                # If paper is None (deleted paper), search for any CandidateAnswer matching candidate+question.
                if paper is not None:
                    ans = CandidateAnswer.objects.filter(candidate=candidate, paper=paper, question=q).first()
                else:
                    ans = CandidateAnswer.objects.filter(candidate=candidate, question=q).first()

                row = [
                    serial,
                    candidate.name,
                    candidate.exam_center,
                    photo_url,
                    father_name,
                    candidate.dob,
                    candidate.rank,
                    candidate.trade.name if getattr(candidate, "trade", None) else "",
                    candidate.army_no,
                    aadhar_number,
                    candidate.primary_qualification,
                    candidate.primary_duration,
                    candidate.primary_credits,
                    candidate.secondary_qualification,
                    candidate.secondary_duration,
                    candidate.secondary_credits,
                    candidate.nsqf_level,
                    getattr(candidate, "training_center", ""),
                    getattr(candidate, "district", ""),
                    getattr(candidate, "state", ""),
                    candidate.primary_viva_marks,
                    # candidate.secondary_viva_marks,
                    candidate.primary_practical_marks,
                    # candidate.secondary_practical_marks,
                    candidate.army_no,
                    "Secondary" if (paper and getattr(paper, "is_common", False)) else ("Primary" if paper else "Unknown"),
                    q.part,
                    q.text,
                    ans.answer if ans and ans.answer is not None else "N/A",
                    getattr(q, "correct_answer", None),
                    q.marks if hasattr(q, "marks") else None,
                ]
                ws.append(row)
                serial += 1

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


# -------------------------
# Crypto helper: encrypt bytes → .dat (salt + iv + ciphertext)
# -------------------------
def _encrypt_bytes_to_dat(data: bytes, passphrase: str) -> bytes:
    if not passphrase:
        raise ValueError("Missing CONVERTER_PASSPHRASE in settings.")

    salt = _os.urandom(16)
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100000,
    )
    key = kdf.derive(passphrase.encode("utf-8"))

    iv = _os.urandom(12)
    aesgcm = AESGCM(key)
    ciphertext = aesgcm.encrypt(iv, data, None)  # AAD=None

    # Layout: salt (16) || iv (12) || ciphertext (includes auth tag)
    return salt + iv + ciphertext


# -------------------------
# DAT exporter (encrypted .xlsx inside, converter-compatible)
# -------------------------
def export_candidates_dat(modeladmin, request, queryset):
    xlsx_bytes = _build_export_workbook(queryset)

    passphrase = getattr(settings, "CONVERTER_PASSPHRASE", None)
    if not passphrase:
        return HttpResponseBadRequest(
            "Server missing CONVERTER_PASSPHRASE; set it in settings or env."
        )

    dat_bytes = _encrypt_bytes_to_dat(xlsx_bytes, passphrase)

    from centers.models import Center

    center = Center.objects.first()

    if center:
        safe_exam_center = "".join(c if c.isalnum() else "_" for c in center.exam_Center)
        safe_comd = "".join(c if c.isalnum() else "_" for c in center.comd)
        filename = f"{safe_comd}_{safe_exam_center}.dat"
    else:
        ts = timezone.now().strftime("%Y%m%d%H%M%S")
        filename = f"candidates_export_{ts}.dat"

    response = HttpResponse(dat_bytes, content_type="application/octet-stream")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# Changed label: this will be displayed as the action/button text
export_candidates_dat.short_description = "Export All Exam Data"


# -------------------------
# Export candidate images as ZIP
# -------------------------
def export_candidate_images(modeladmin, request, queryset):
    from io import BytesIO

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for candidate in queryset:
            photo = getattr(candidate, "photograph", None)
            if photo:
                try:
                    file_path = photo.path
                    ext = file_path[file_path.rfind(".") :] if "." in file_path else ""
                    filename = f"{candidate.army_no}_{candidate.name}{ext}"
                    zip_file.write(file_path, arcname=filename)
                except Exception:
                    continue

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="candidate_images.zip"'
    return response


# changed label as requested
export_candidate_images.short_description = "Export All Photos"


def export_all_candidate_images(modeladmin, request):
    from io import BytesIO

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for candidate in CandidateProfile.objects.all():
            photo = getattr(candidate, "photograph", None)
            if photo:
                try:
                    file_path = photo.path
                    ext = file_path[file_path.rfind(".") :] if "." in file_path else ""
                    filename = f"{candidate.army_no}_{candidate.name}{ext}"
                    zip_file.write(file_path, arcname=filename)
                except Exception:
                    continue

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="all_candidate_images.zip"'
    return response


# -------------------------
# NEW: Export marks (primary/secondary viva & practical) to Excel for selected queryset
# -------------------------
def export_marks_excel(modeladmin, request, queryset):
    """
    Export a simple Excel sheet with marks columns for the selected candidates.
    Columns: Army No, Name, Trade, Primary Viva, Primary Practical, Training Center, Exam Center, Created At
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks"

    columns = [
        "Army No",
        "Name",
        "Trade",
        "Primary Viva Marks",
        "Primary Practical Marks",
        # "Secondary Viva Marks",
        # "Secondary Practical Marks",
        "Training Center",
        "Exam Center",
        "Created At",
    ]

    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num).value = column_title

    for row_num, candidate in enumerate(queryset, 2):
        trade_obj = getattr(candidate, "trade", None)
        trade_name = getattr(trade_obj, "name", str(trade_obj)) if trade_obj else ""
        row = [
            candidate.army_no,
            candidate.name,
            trade_name,
            candidate.primary_viva_marks,
            candidate.primary_practical_marks,
            # candidate.secondary_viva_marks,
            # candidate.secondary_practical_marks,
            getattr(candidate, "training_center", ""),
            candidate.exam_center,
            candidate.created_at.strftime("%Y-%m-%d %H:%M") if candidate.created_at else "",
        ]
        for col_num, cell_value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="candidate_marks.xlsx"'
    wb.save(response)
    return response


export_marks_excel.short_description = "Export Viva-Prac Marks"


# -------------------------
# Admin Registration
# -------------------------
@admin.register(CandidateProfile)
class CandidateProfileAdmin(admin.ModelAdmin):
    form = CandidateProfileAdminForm

    # default (non-PO) list
    list_display = ("army_no", "name", "user", "rank", "trade", "shift", "created_at")
    # base declaration; we will set this per-request in changelist_view
    list_editable = ()
    list_filter = ("trade", "training_center")

    # all actions declared; we'll filter them per user in get_actions
    actions = [
        export_candidate_answers,
        export_candidates_excel,
        export_candidates_dat,
        export_candidate_images,
        export_marks_excel,  # include new marks export as an action
    ]

    # ---------- helpers ----------
    def _is_po(self, request):
        """Return True if user is PO by group or (optional) role."""
        u = request.user
        in_po_group = u.groups.filter(name="PO").exists()
        has_po_role = getattr(u, "role", None) == "PO_ADMIN"
        return in_po_group or has_po_role

    def _field_exists(self, field_name: str) -> bool:
        """Check if a given field actually exists on CandidateProfile."""
        return any(f.name == field_name for f in CandidateProfile._meta.get_fields())

    # ---------- changelist (top buttons/links area) ----------
    def changelist_view(self, request, extra_context=None):
        # Turn on inline editing only for PO (important: set attribute here)
        if self._is_po(request):
            self.list_editable = (
                "primary_viva_marks",
                "primary_practical_marks",
                # "secondary_viva_marks",
                # "secondary_practical_marks",
            )
        else:
            self.list_editable = ()  # no inline editing for others

        # IMPORTANT: do not inject export_all_* keys here (we remove top links client-side)
        return super().changelist_view(request, extra_context=extra_context)

    # ---------- list columns ----------
    def get_list_display(self, request):
        if self._is_po(request):
            # Minimal columns for PO with marks
            return (
                "army_no",
                "name",
                "trade",
                "primary_viva_marks",
                "primary_practical_marks",
                # "secondary_viva_marks",
                # "secondary_practical_marks",
            )
        # non-PO (your original default)
        return ("army_no", "name", "user", "rank", "trade", "shift", "created_at")

    # Remove row links for PO
    def get_list_display_links(self, request, list_display):
        if self._is_po(request):
            return ()  # no links → can't open detail page
        return super().get_list_display_links(request, list_display)

    # ---------- actions per role ----------
    def get_actions(self, request):
        actions = super().get_actions(request)
        if self._is_po(request):
            # PO can export DAT, Photos and Marks; hide other actions if you prefer
            return {
                k: v
                for k, v in actions.items()
                if k in ["export_candidates_dat", "export_candidate_images", "export_marks_excel"]
            }
        else:
            # Non-PO cannot export DAT, Photos or Marks via actions
            blocked = {"export_candidates_dat", "export_candidate_images", "export_marks_excel"}
            return {k: v for k, v in actions.items() if k not in blocked}

    # ---------- change form (add/change page) ----------
    def get_fields(self, request, obj=None):
        # Base fields for a full candidate (non-PO users)
        raw_base_fields = [
            "user",
            "army_no",
            "rank",
            "trade_type",
            "unit",
            "brigade",
            "corps",
            "command",
            "trade",
            "name",
            "dob",
            "doe",
            "aadhar_number",
            "father_name",
            "photograph",
            "med_cat",
            "cat",
            "nsqf_level",
            "exam_center",
            "training_center",
            "state",
            "district",
            "primary_qualification",
            "primary_duration",
            "primary_credits",
            # "secondary_qualification",
            # "secondary_duration",
            # "secondary_credits",
            "shift",
        ]
        base_fields = [f for f in raw_base_fields if self._field_exists(f)]

        po_only_fields_raw = [
            "primary_viva_marks",
            "primary_practical_marks",
            "secondary_viva_marks",
            "secondary_practical_marks",
        ]
        po_only_fields = [f for f in po_only_fields_raw if self._field_exists(f)]

        if self._is_po(request):
            # If someone hits the change URL directly, still show limited fields
            po_base = [f for f in ["army_no", "name", "trade"] if self._field_exists(f)]
            return po_base + po_only_fields

        # Non-PO: base fields + created_at if it exists
        if self._field_exists("created_at"):
            return base_fields + ["created_at"]
        return base_fields

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if self._is_po(request):
            # PO can edit only the four marks; everything else readonly
            all_possible = [
                "user",
                "army_no",
                "rank",
                "trade_type",
                "unit",
                "brigade",
                "corps",
                "command",
                "trade",
                "name",
                "dob",
                "doe",
                "aadhar_number",
                "father_name",
                "photograph",
                "med_cat",
                "cat",
                "nsqf_level",
                "exam_center",
                "training_center",
                "state",
                "district",
                "primary_qualification",
                "primary_duration",
                "primary_credits",
                # "secondary_qualification",
                # "secondary_duration",
                # "secondary_credits",
                "shift",
                "created_at",
            ]
            for f in all_possible:
                if self._field_exists(f) and f not in readonly:
                    readonly.append(f)
        else:
            if self._field_exists("created_at") and "created_at" not in readonly:
                readonly.append("created_at")
        return readonly

    # Block opening the change form UI for PO (must use list editing)
    def change_view(self, request, object_id, form_url="", extra_context=None):
        if self._is_po(request):
            return HttpResponseForbidden("PO edits marks on the list page only.")
        return super().change_view(request, object_id, form_url, extra_context)

    # prevent PO from add/delete
    def has_add_permission(self, request):
        if self._is_po(request):
            return False
        return super().has_add_permission(request)

    def has_delete_permission(self, request, obj=None):
        if self._is_po(request):
            return False
        return super().has_delete_permission(request, obj)

    # Optional per-object link (not used by PO because links are disabled)
    def download_csv_link(self, obj):
        url = reverse("export_candidate_pdf", args=[obj.id])
        return format_html('<a class="button" href="{}">Download Answers PDF</a>', url)

    download_csv_link.short_description = "Export PDF"
    download_csv_link.allow_tags = True

    # ---------- Custom URLs (only PO allowed) ----------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "Export-All-dat/",
                self.admin_site.admin_view(self.export_all_dat_view),
                name="registration_candidateprofile_export_all_dat",
            ),
            path(
                "Export-All-Images/",
                self.admin_site.admin_view(self.export_all_images_view),
                name="registration_candidateprofile_export_all_images",
            ),
            # NEW: Export-All-Marks (admin-bound method)
            path(
                "Export-All-Marks/",
                self.admin_site.admin_view(self.export_all_marks_view),
                name="registration_candidateprofile_export_all_marks",
            ),
            # JS endpoint that injects the sidebar buttons (served via admin view to allow permission check)
            path(
                "candidate-export-links.js",
                self.admin_site.admin_view(self.export_links_js),
                name="registration_candidateprofile_export_links_js",
            ),
        ]
        return custom_urls + urls

    def export_all_dat_view(self, request):
        # Only PO can export DAT
        if not self._is_po(request):
            return HttpResponseForbidden("Not allowed.")
        qs = self.get_queryset(request)
        return export_candidates_dat(self, request, qs)

    def export_all_images_view(self, request):
        # Only PO can export photos ZIP
        if not self._is_po(request):
            return HttpResponseForbidden("Not allowed.")
        return export_all_candidate_images(self, request)

    def export_all_marks_view(self, request):
        # Only PO can export Marks
        if not self._is_po(request):
            return HttpResponseForbidden("Not allowed.")
        qs = self.get_queryset(request)
        return export_marks_excel(self, request, qs)

    def export_links_js(self, request):
        """
        Serve a small JS file that:
          - fixes header checkbox / row checkbox behavior (for all admin users)
          - adds basic tooltips (title attributes) to sidebar links (for all)
          - for PO users only: removes header anchors to export endpoints and inserts sidebar export buttons (DAT / Images / Marks)
        """
        # Determine if current user is PO
        is_po = self._is_po(request)

        # Build URLs & labels (used only when is_po)
        dat_url = reverse("admin:registration_candidateprofile_export_all_dat")
        img_url = reverse("admin:registration_candidateprofile_export_all_images")
        marks_url = reverse("admin:registration_candidateprofile_export_all_marks")

        dat_label = export_candidates_dat.short_description or "Export All Answers"
        img_label = export_candidate_images.short_description or "Export All Photos"
        marks_label = export_marks_excel.short_description or "Export All Marks"

        candidate_changelist = reverse("admin:registration_candidateprofile_changelist")

        # JS template:
        js_template = r"""
(function(){
    try {
        function onReady(fn) {
            if (document.readyState !== 'loading') {
                fn();
            } else {
                document.addEventListener('DOMContentLoaded', fn);
            }
        }

        onReady(function() {

            /********** 1) FIX header / row checkboxes (run for all users) **********/
            try {
                var rowCheckboxes = Array.prototype.slice.call(document.querySelectorAll('input.action-select, input[name="_selected_action"]'));
                if (rowCheckboxes.length) {
                    rowCheckboxes.forEach(function(cb) {
                        var style = window.getComputedStyle(cb);
                        if (style && style.display !== 'none') {
                            cb.checked = true;
                            cb.dispatchEvent(new Event('change', {bubbles:true}));
                        }
                    });
                    var master = document.getElementById('action-toggle') || document.querySelector('thead input[type="checkbox"]');
                    if (master) {
                        var allChecked = rowCheckboxes.every(function(c){ return c.checked; });
                        master.checked = allChecked;
                        master.dispatchEvent(new Event('change', {bubbles:true}));
                        master.addEventListener('click', function() {
                            var checked = !!master.checked;
                            rowCheckboxes.forEach(function(c) {
                                c.checked = checked;
                                c.dispatchEvent(new Event('change', {bubbles:true}));
                            });
                        });
                    } else {
                        var headerLabel = document.querySelector('thead th .action-checkbox, thead th .action-select');
                        if (headerLabel) {
                            headerLabel.addEventListener('click', function() {
                                var anyUnchecked = rowCheckboxes.some(function(c){ return !c.checked; });
                                var newState = anyUnchecked;
                                rowCheckboxes.forEach(function(c) {
                                    c.checked = newState;
                                    c.dispatchEvent(new Event('change', {bubbles:true}));
                                });
                            });
                        }
                    }
                }
            } catch (e) {
                console.error('selection-fix error', e);
            }

            /********** 2) Add basic tooltips to sidebar links (run for all users) **********/
            try {
                var sidebarAnchors = Array.prototype.slice.call(document.querySelectorAll('#sidebar a, .app-list a'));
                sidebarAnchors.forEach(function(a) {
                    try {
                        var txt = (a.textContent || a.innerText || '').trim();
                        if (txt && !a.getAttribute('title')) {
                            a.setAttribute('title', txt);
                        }
                    } catch(e) { }
                });
            } catch (e) {
                console.error('sidebar-tooltip error', e);
            }

            /********** 3) PO-only actions: remove header export links and inject sidebar buttons **********/
            var isPo = {IS_PO};
            if (isPo) {
                try {
                    var removeSelectors = ['a[href*="{DAT_URL}"]','a[href*="{IMG_URL}"]','a[href*="{MARKS_URL}"]'];
                    removeSelectors.forEach(function(sel) {
                        var nodes = document.querySelectorAll(sel);
                        nodes.forEach(function(n) {
                            if (n && n.parentNode) n.parentNode.removeChild(n);
                        });
                    });
                } catch (e) {}

                try {
                    var targetHref = "{CANDIDATE_CHANGELIST}";
                    var anchors = Array.prototype.slice.call(document.querySelectorAll('a'));
                    var targetAnchor = null;
                    for (var i=0;i<anchors.length;i++) {
                        var a = anchors[i];
                        try { var href = a.getAttribute('href') || ''; } catch(e) { var href = ''; }
                        if (!href) continue;
                        var normalized = href.replace(window.location.origin, '');
                        if (normalized.indexOf(targetHref) !== -1 || normalized === targetHref) {
                            targetAnchor = a;
                            break;
                        }
                    }
                    if (targetAnchor) {
                        var parentLi = targetAnchor.closest('li');
                        var insertAfter = parentLi || targetAnchor;

                        // createButton now accepts tooltip text
                        function createButton(href, label, tooltip) {
                            var a = document.createElement('a');
                            a.setAttribute('href', href);
                            a.setAttribute('title', tooltip || label);
                            a.setAttribute('aria-label', tooltip || label);
                            a.style.display = 'block';
                            a.style.padding = '6px 10px';
                            a.style.marginTop = '4px';
                            a.style.marginLeft = parentLi ? '18px' : '10px';
                            a.style.borderRadius = '4px';
                            a.style.background = 'transparent';
                            a.style.color = '#fff';
                            a.style.textDecoration = 'none';
                            a.innerText = label;
                            return a;
                        }

                        var wrapper = document.createElement('div');
                        wrapper.className = 'candidate-export-buttons';
                        wrapper.style.padding = '4px 0 6px 0';

                        // NOTE: tooltip strings below are what will show on hover
                        var b1 = createButton("{DAT_URL}", "{DAT_LABEL}", "Download encrypted .dat (for Converter)");
                        var b2 = createButton("{IMG_URL}", "{IMG_LABEL}", "Download ZIP of all candidate photos");
                        var b3 = createButton("{MARKS_URL}", "{MARKS_LABEL}", "Download Excel with Viva & Practical marks");

                        b1.style.fontWeight = '600'; b2.style.fontWeight = '600'; b3.style.fontWeight = '600';

                        wrapper.appendChild(b1);
                        wrapper.appendChild(b2);
                        wrapper.appendChild(b3);

                        if (insertAfter && insertAfter.parentNode) {
                            if (insertAfter.nextSibling) {
                                insertAfter.parentNode.insertBefore(wrapper, insertAfter.nextSibling);
                            } else {
                                insertAfter.parentNode.appendChild(wrapper);
                            }
                        } else {
                            var sidebar = document.querySelector('#sidebar') || document.querySelector('.module');
                            if (sidebar) sidebar.appendChild(wrapper);
                        }
                    }
                } catch (e) {
                    console.error('po-sidebar-insert error', e);
                }
            } // end isPo

        }); // onReady
    } catch (e) {
        console.error('export_links_js top error', e);
    }
})();
"""

        # Safe substitutions
        js = js_template.replace("{IS_PO}", "true" if is_po else "false")
        js = js.replace("{DAT_URL}", dat_url)
        js = js.replace("{IMG_URL}", img_url)
        js = js.replace("{MARKS_URL}", marks_url)
        js = js.replace("{DAT_LABEL}", dat_label.replace('"', '\\"'))
        js = js.replace("{IMG_LABEL}", img_label.replace('"', '\\"'))
        js = js.replace("{MARKS_LABEL}", marks_label.replace('"', '\\"'))
        js = js.replace("{CANDIDATE_CHANGELIST}", candidate_changelist)

        return HttpResponse(js, content_type="application/javascript")

    # Ensure the admin loads our small JS file (served by export_links_js)
    @property
    def media(self):
        try:
            js_url = reverse("admin:registration_candidateprofile_export_links_js")
            return forms.Media(js=(js_url,))
        except Exception:
            # If reverse fails for any reason, return empty Media to avoid breaking admin
            return forms.Media()